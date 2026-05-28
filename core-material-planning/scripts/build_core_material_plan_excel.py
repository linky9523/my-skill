#!/usr/bin/env python3
"""Build a smartphone core material planning workbook from JSON."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MAIN_HEADERS = ["素材分类", "需求细则", "需求说明", "手机展现", "需求数量", "优先级"]
CONFIRM_HEADERS = ["待确认项", "原因", "建议确认方式", "关联素材"]

COLORS = {
    "title": "046A38",
    "header": "DDF2E6",
    "group": "EAF8EF",
    "subtle": "F7FBF8",
    "white": "FFFFFF",
    "line": "B7D8C4",
    "priority": "FFF0C2",
    "todo": "FFF7DA",
    "ink": "24312A",
    "muted": "5B6B60",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "、".join(text(item) for item in value if text(item))
    return str(value).strip()


def load_data(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise ValueError("Input JSON must be an object.")
    return data


def normalize_material(row: dict[str, Any]) -> list[str]:
    return [
        text(row.get("category")),
        text(row.get("detail")),
        text(row.get("description")),
        text(row.get("phone_display")),
        text(row.get("quantity")),
        text(row.get("priority")),
    ]


def normalize_confirmation(row: dict[str, Any]) -> list[str]:
    return [
        text(row.get("item")),
        text(row.get("reason")),
        text(row.get("suggestion")),
        text(row.get("related_material")),
    ]


def thin_border() -> Border:
    side = Side(style="thin", color=COLORS["line"])
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_row(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        cell.font = Font(name="Arial", size=11, bold=True, color=COLORS["ink"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def style_body(ws, min_row: int, max_row: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        fill = COLORS["white"] if row % 2 else COLORS["subtle"]
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Arial", size=10, color=COLORS["ink"])
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border()
        priority = text(ws.cell(row, max_col).value)
        if priority:
            ws.cell(row, max_col).fill = PatternFill("solid", fgColor=COLORS["priority"])
            ws.cell(row, max_col).font = Font(name="Arial", size=10, bold=True, color=COLORS["ink"])


def merge_span(ws, start_row: int, end_row: int, col: int) -> None:
    if end_row <= start_row:
        return
    value = text(ws.cell(start_row, col).value)
    if not value:
        return
    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
    cell = ws.cell(start_row, col)
    cell.fill = PatternFill("solid", fgColor=COLORS["group"])
    cell.font = Font(name="Arial", size=10, bold=True, color=COLORS["ink"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def merge_main_groups(ws, first_row: int, last_row: int) -> None:
    row = first_row
    while row <= last_row:
        category = text(ws.cell(row, 1).value)
        cat_end = row
        while cat_end + 1 <= last_row and text(ws.cell(cat_end + 1, 1).value) == category:
            cat_end += 1

        detail_row = row
        while detail_row <= cat_end:
            detail = text(ws.cell(detail_row, 2).value)
            detail_end = detail_row
            while detail_end + 1 <= cat_end and text(ws.cell(detail_end + 1, 2).value) == detail:
                detail_end += 1
            merge_span(ws, detail_row, detail_end, 2)
            detail_row = detail_end + 1

        merge_span(ws, row, cat_end, 1)
        row = cat_end + 1


def setup_sheet(ws, title: str, subtitle: str, headers: list[str], rows: list[list[str]], merge_groups: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    max_col = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = Font(name="Arial", size=16, bold=True, color=COLORS["white"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["title"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sub_cell = ws.cell(2, 1, subtitle)
    sub_cell.font = Font(name="Arial", size=10, color=COLORS["muted"])
    sub_cell.fill = PatternFill("solid", fgColor=COLORS["subtle"])
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, start=1):
        ws.cell(3, col, header)
    style_header_row(ws, 3, max_col)

    for row_idx, row_values in enumerate(rows, start=4):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row_idx, col_idx, value)

    last_row = max(4, 3 + len(rows))
    if rows:
        style_body(ws, 4, last_row, max_col)
        ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{last_row}"
        if merge_groups:
            merge_main_groups(ws, 4, last_row)

    widths = [18, 20, 46, 18, 12, 16] if max_col == 6 else [24, 42, 42, 32]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 26
    for row in range(4, last_row + 1):
        ws.row_dimensions[row].height = 42


def build_workbook(data: dict[str, Any]) -> Workbook:
    product = data.get("product") or {}
    product_name = text(product.get("name")) or "手机产品"
    positioning = text(product.get("positioning"))
    colors = text(product.get("colors"))
    source = text(product.get("source"))

    materials = data.get("materials") or []
    if not isinstance(materials, list) or not materials:
        raise ValueError("Input JSON must include a non-empty materials list.")
    material_rows = [normalize_material(row) for row in materials]

    confirmations = data.get("confirmations") or []
    if confirmations and not isinstance(confirmations, list):
        raise ValueError("confirmations must be a list when provided.")
    confirmation_rows = [normalize_confirmation(row) for row in confirmations]
    if not confirmation_rows:
        confirmation_rows = [["暂无", "已在主清单中完成基础规划", "", ""]]

    subtitle_parts = []
    if positioning:
        subtitle_parts.append(positioning)
    if colors:
        subtitle_parts.append(f"颜色：{colors}")
    if source:
        subtitle_parts.append(f"来源：{source}")
    subtitle = " | ".join(subtitle_parts)

    wb = Workbook()
    ws = wb.active
    ws.title = "核心素材清单"
    setup_sheet(ws, f"{product_name} 核心素材清单", subtitle, MAIN_HEADERS, material_rows, merge_groups=True)

    todo = wb.create_sheet("待确认项")
    setup_sheet(todo, f"{product_name} 待确认项", "记录官网/PRD未给出的执行细节，不在主清单中编造。", CONFIRM_HEADERS, confirmation_rows)

    return wb


def validate_output(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "核心素材清单" not in wb.sheetnames or "待确认项" not in wb.sheetnames:
        raise ValueError("Workbook must include 核心素材清单 and 待确认项 sheets.")
    ws = wb["核心素材清单"]
    headers = [text(ws.cell(3, col).value) for col in range(1, 7)]
    if headers != MAIN_HEADERS:
        raise ValueError(f"Main headers mismatch: {headers}")
    all_headers = [text(ws.cell(3, col).value) for col in range(1, ws.max_column + 1)]
    if "参考" in all_headers:
        raise ValueError("Main sheet must not include 参考 column.")


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("Usage: build_core_material_plan_excel.py input.json output.xlsx", file=sys.stderr)
        return 2

    input_path = Path(argv[1])
    output_path = Path(argv[2])
    data = load_data(input_path)
    wb = build_workbook(data)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    validate_output(output_path)
    print(f"Saved {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
