#!/usr/bin/env python3
"""Build a poster-like product infographic Excel wireframe from JSON."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COLS = 16
DEFAULT_POSTER_ROWS = 128
MAX_POSTER_MODULES = 4
MAX_POSTER_POINTS_PER_MODULE = 3
EXCLUDED_TERMS = ("价格", "售价", "开售", "发售", "预售", "预约", "渠道", "购买", "二维码", "套装", "sku", "SKU")

COLORS = {
    "paper": "F6F8FA",
    "white": "FFFFFF",
    "ink": "111827",
    "muted": "6B7280",
    "line": "D8DEE9",
    "dark": "101820",
    "dark2": "1F2937",
    "cyan": "E7FBF8",
    "orange": "FFF4D6",
    "blue": "EAF2FF",
    "green": "EAFBF0",
    "gray": "EEF2F7",
    "red": "FEE2E2",
}


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(text(item) for item in value if text(item))
    return str(value)


def as_list(value: Any) -> list[Any]:
    if not value:
        return []
    return value if isinstance(value, list) else [value]


def is_excluded_module(module: dict[str, Any]) -> bool:
    haystack = " ".join(
        text(module.get(k, "")) for k in ("section", "section_title", "subtitle", "purpose", "layout", "visual", "notes")
    )
    return any(term in haystack for term in EXCLUDED_TERMS)


def side(color: str = "CDD6E1", style: str = "thin") -> Side:
    return Side(style=style, color=color)


def simple_border(color: str = "CDD6E1", style: str = "thin") -> Border:
    s = side(color, style)
    return Border(left=s, right=s, top=s, bottom=s)


def fill_area(ws, r1: int, c1: int, r2: int, c2: int, fill: str) -> None:
    for row in ws.iter_rows(min_row=r1, min_col=c1, max_row=r2, max_col=c2):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", size=10, color=COLORS["ink"])


def outline_range(ws, r1: int, c1: int, r2: int, c2: int, color: str = "CDD6E1", style: str = "medium") -> None:
    top = side(color, style)
    bottom = side(color, style)
    left = side(color, style)
    right = side(color, style)
    for c in range(c1, c2 + 1):
        ws.cell(r1, c).border = Border(top=top)
        ws.cell(r2, c).border = Border(bottom=bottom)
    for r in range(r1, r2 + 1):
        ws.cell(r, c1).border = Border(left=left)
        ws.cell(r, c2).border = Border(right=right)
    ws.cell(r1, c1).border = Border(top=top, left=left)
    ws.cell(r1, c2).border = Border(top=top, right=right)
    ws.cell(r2, c1).border = Border(bottom=bottom, left=left)
    ws.cell(r2, c2).border = Border(bottom=bottom, right=right)


def merge_box(
    ws,
    r1: int,
    c1: int,
    r2: int,
    c2: int,
    value: str,
    fill: str = "FFFFFF",
    font_color: str = "111827",
    size: int = 12,
    bold: bool = False,
    align: str = "center",
    valign: str = "center",
    outline: str = "D8DEE9",
    thick: bool = False,
) -> None:
    fill_area(ws, r1, c1, r2, c2, fill)
    for row in ws.iter_rows(min_row=r1, min_col=c1, max_row=r2, max_col=c2):
        for cell in row:
            cell.border = simple_border(outline, "medium" if thick else "thin")
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1, value)
    cell.font = Font(name="Arial", size=size, color=font_color, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=True)


def setup_poster_sheet(ws, rows: int = DEFAULT_POSTER_ROWS) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 65
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    for col in range(1, COLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8.7
    for row in range(1, rows + 1):
        ws.row_dimensions[row].height = 22
        for col in range(1, COLS + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=COLORS["paper"])
            cell.border = Border()


def add_card(ws, r1: int, c1: int, r2: int, c2: int, point: dict[str, Any], accent: str = "blue") -> None:
    fill_area(ws, r1, c1, r2, c2, COLORS["white"])
    title = text(point.get("title", ""))
    benefit = text(point.get("benefit", ""))
    proof = " | ".join(part for part in [text(point.get("proof", "")), text(point.get("parameters", ""))] if part)
    visual = text(point.get("visual", ""))
    height = r2 - r1 + 1

    merge_box(ws, r1, c1, r1, min(c2, c1 + 2), "文案", fill=COLORS.get(accent, COLORS["gray"]), size=9, bold=True)
    if height <= 5:
        merge_box(ws, r1 + 1, c1, r1 + 1, c2, title, fill=COLORS["white"], size=12, bold=True)
        if benefit:
            merge_box(ws, r1 + 2, c1, r1 + 2, c2, benefit, fill=COLORS["white"], font_color=COLORS["muted"], size=9)
        if visual:
            merge_box(ws, r1 + 3, c1, r2, c2, f"图片占位：{visual}", fill=COLORS.get(accent, COLORS["gray"]), font_color=COLORS["muted"], size=9)
    else:
        title_end = min(r1 + 2, r2 - 3)
        merge_box(ws, r1 + 1, c1, title_end, c2, title, fill=COLORS["white"], size=13, bold=True)
        if benefit:
            merge_box(ws, title_end + 1, c1, title_end + 1, c2, benefit, fill=COLORS["white"], font_color=COLORS["muted"], size=10)
        if proof and title_end + 2 <= r2 - 3:
            merge_box(ws, title_end + 2, c1, title_end + 2, c2, proof, fill=COLORS["white"], size=8)
        if visual:
            merge_box(ws, r2 - 2, c1, r2, c2, f"图片占位：{visual}", fill=COLORS.get(accent, COLORS["gray"]), font_color=COLORS["muted"], size=10)
    outline_range(ws, r1, c1, r2, c2)


def add_kv(ws, row: int, data: dict[str, Any], module: dict[str, Any] | None) -> int:
    product = data.get("product", {})
    name = text(product.get("name") or (module or {}).get("section_title") or "产品名")
    slogan = text(product.get("slogan") or (module or {}).get("subtitle") or product.get("positioning") or "")
    visual = text((module or {}).get("visual") or "产品 KV / 产品正反面 / 品牌或联合研发标识 / 主视觉背景")
    note = text((module or {}).get("notes") or "这一屏只做产品心智，不塞卖点卡片")

    merge_box(ws, row, 2, row + 2, 15, name, fill=COLORS["paper"], size=25, bold=True)
    if slogan:
        merge_box(ws, row + 3, 5, row + 3, 12, slogan, fill=COLORS["paper"], font_color=COLORS["dark2"], size=14, bold=True)
    merge_box(
        ws,
        row + 6,
        3,
        row + 16,
        14,
        f"产品 KV 大图占位\n{visual}\n\n设计提示：{note}",
        fill=COLORS["white"],
        font_color=COLORS["muted"],
        size=15,
        bold=True,
        outline="BFD7D4",
        thick=True,
    )
    return row + 19


def add_section_title(ws, row: int, title: str, subtitle: str = "") -> int:
    merge_box(ws, row, 2, row + 1, 15, title, fill=COLORS["paper"], size=18, bold=True)
    if subtitle:
        merge_box(ws, row + 2, 4, row + 2, 13, subtitle, fill=COLORS["paper"], font_color=COLORS["muted"], size=10)
        return row + 4
    return row + 3


def add_module(ws, row: int, module: dict[str, Any], idx: int) -> int:
    title = text(module.get("section_title") or module.get("section") or f"模块 {idx}")
    subtitle = text(module.get("subtitle") or module.get("purpose") or "")
    layout = text(module.get("layout"))
    visual = text(module.get("visual"))
    points = as_list(module.get("selling_points"))[:MAX_POSTER_POINTS_PER_MODULE]
    row = add_section_title(ws, row, title, subtitle)

    if points:
        lead = points[0]
        if len(points) == 1:
            add_card(ws, row, 2, row + 8, 15, lead, "blue")
            row += 10
        else:
            add_card(ws, row, 2, row + 10, 9, lead, "blue")
            for i, point in enumerate(points[1:3]):
                r1 = row + i * 5
                add_card(ws, r1, 10, r1 + 4, 15, point, "cyan" if i == 0 else "green")
            row += 12

    copy_lines = [text(block.get("text", "")) for block in as_list(module.get("copy_blocks")) if text(block.get("text", ""))]
    if layout or visual or copy_lines:
        parts = []
        if copy_lines:
            parts.append("补充文案：\n" + "\n".join(copy_lines))
        if layout:
            parts.append("版式建议：" + layout)
        if visual:
            parts.append("素材需求：" + visual)
        merge_box(ws, row, 2, row + 2, 15, "\n".join(parts), fill=COLORS["cyan"], font_color=COLORS["dark2"], size=10, bold=True, outline="BFD7D4", thick=True)
        row += 4

    return row + 1


def add_poster_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "海报式排版"
    setup_poster_sheet(ws)
    modules = [m for m in sorted(as_list(data.get("modules")), key=lambda x: x.get("order", 999)) if not is_excluded_module(m)]
    kv = next((m for m in modules if text(m.get("section", "")).lower() == "kv"), None)
    modules = [m for m in modules if m is not kv][:MAX_POSTER_MODULES]

    row = add_kv(ws, 2, data, kv)
    for idx, module in enumerate(modules, start=1):
        row = add_module(ws, row, module, idx)

    merge_box(
        ws,
        row,
        2,
        row + 1,
        15,
        "给设计师的读法：这是压缩版一图读懂，主图只放核心卡片；完整卖点、参数和素材需求见后面的清单。",
        fill=COLORS["orange"],
        font_color=COLORS["dark2"],
        size=11,
        bold=True,
    )
    ws.print_area = f"A1:P{min(max(row + 1, 1), DEFAULT_POSTER_ROWS)}"


def add_copy_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("文案与素材清单")
    ws.sheet_view.showGridLines = False
    headers = ["顺序", "模块", "主文案", "利益点", "参数/证明", "素材需求", "来源", "备注"]
    ws.append(headers)
    modules = [m for m in sorted(as_list(data.get("modules")), key=lambda x: x.get("order", 999)) if not is_excluded_module(m)]
    order = 1
    for module in modules:
        section = text(module.get("section") or module.get("section_title"))
        if module.get("section_title"):
            ws.append([order, section, text(module.get("section_title")), text(module.get("subtitle") or module.get("purpose")), "", text(module.get("visual")), "", text(module.get("notes"))])
            order += 1
        for point in as_list(module.get("selling_points")):
            proof = " | ".join(part for part in [text(point.get("proof")), text(point.get("parameters"))] if part)
            ws.append([order, section, text(point.get("title")), text(point.get("benefit")), proof, text(point.get("visual")), text(point.get("source")), text(point.get("notes"))])
            order += 1

    widths = [8, 16, 30, 28, 36, 32, 22, 24]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    style_table(ws)


def add_questions_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("待确认项")
    ws.sheet_view.showGridLines = False
    ws.append(["事项", "为什么要确认", "建议负责人", "建议动作", "状态"])
    for q in as_list(data.get("open_questions")):
        item = text(q.get("item", ""))
        if any(term in item for term in EXCLUDED_TERMS):
            continue
        ws.append([item, text(q.get("reason", "")), text(q.get("owner", "")), text(q.get("suggested_action", "")), text(q.get("status", "待确认"))])
    if ws.max_row == 1:
        ws.append(["暂无", "源文件中暂未发现需要单独确认的产品策划项", "", "", ""])
    widths = [26, 52, 18, 36, 14]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    style_table(ws)


def style_table(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = simple_border()
            cell.font = Font(name="Arial", size=10, color=COLORS["ink"])
            cell.fill = PatternFill("solid", fgColor=COLORS["white"])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=COLORS["dark"])
        cell.font = Font(name="Arial", size=10, color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def build(data: dict[str, Any], output: Path) -> None:
    wb = Workbook()
    add_poster_sheet(wb, data)
    add_copy_sheet(wb, data)
    add_questions_sheet(wb, data)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("Usage: build_infographic_excel.py input.json output.xlsx", file=sys.stderr)
        return 2
    build(load_json(Path(argv[1])), Path(argv[2]))
    print(argv[2])
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
