#!/usr/bin/env python3
"""Generate rightmost one-to-one official website remarks for workbook rows."""

from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_KEYS = {"章节", "卖点", "文案", "参数"}
EXISTING_REMARK_HEADERS = {"备注", "文案备注", "备注说明", "remark", "remarks", "note", "notes"}
OUTPUT_HEADER = "备注生成"


ZH_TEMPLATES = {
    "general": "产品图片仅供参考，请以实物为准。产品规格、外观细节、材质及工艺可能因生产批次、供应商、测量方法等因素略有差异，请以产品实际为准。",
    "display": "屏幕相关数据来源于 OPPO 实验室，实际显示效果可能因使用环境、设备状态、软件版本、测试方法及个人使用习惯不同而存在差异，请以实际体验为准。护眼相关功能非医疗器械功能，不具有治疗作用。",
    "camera": "影像规格、焦段、进光量、动态范围等数据来源于 OPPO 实验室或供应商测试。实际拍摄效果可能因拍摄环境、拍摄对象、设备状态、软件版本及拍摄模式不同而存在差异，请以实际体验为准。",
    "sample": "页面样张、视频、产品界面、结构示意图及素材效果仅供参考，可能经过裁切、压缩、合成或创意化呈现，请以实际拍摄效果、产品界面和最终上线页面为准。",
    "zoom": "变焦、等效焦段及远摄效果可能因拍摄距离、光线、模式、算法处理及设备状态不同而存在差异。光学品质变焦、数字变焦等能力请以产品实际支持情况为准。",
    "video": "视频、HDR、Log、LUT、实况照片等功能的可用规格和效果与拍摄模式、软件版本、存储空间、环境光及设备状态有关，请以实际体验为准。",
    "software": "以上软件界面和功能仅作展示。AI 及系统相关能力可能因软件版本、地区、账号登录状态、网络环境、适配应用及个人使用习惯不同而存在差异，请以实际版本功能及界面为准。",
    "battery": "电池容量、续航和充电数据来源于 OPPO 实验室或供应商测试，实际表现可能因网络环境、使用方式、充电条件、设备温度、电池老化及其他因素不同而存在差异，请以实际体验为准。",
    "durability": "防水、防尘、抗摔、抗划等数据来源于受控实验室条件。相关防护能力并非永久有效，可能因日常磨损而下降；请勿在潮湿状态下充电，因进液、跌落、碰撞等导致的损坏请以保修政策为准。",
    "performance": "性能、跑分、存储、散热及游戏表现数据来源于 OPPO 实验室或官方测试，实际表现可能因测试环境、软件版本、内存规格、网络环境、应用适配及个人使用习惯不同而存在差异。",
    "connectivity": "通信、音频、协同及第三方服务的功能表现可能因地区、运营商、网络环境、设备状态、软件版本、应用适配及第三方服务政策不同而存在差异，请以实际支持情况为准。",
}

EN_TEMPLATES = {
    "general": "Product images and descriptions are for reference only. Specifications, appearance, materials and craftsmanship may vary due to production batches, suppliers or measurement methods. Please refer to the actual product.",
    "display": "Display data is based on OPPO laboratory tests. Actual display performance may vary depending on environment, device status, software version, test method and usage habits. Eye-care features are not medical functions.",
    "camera": "Camera specifications, focal lengths, light intake and dynamic range data are based on OPPO laboratory or supplier tests. Actual results may vary depending on shooting conditions, subject, device status, software version and mode.",
    "sample": "Sample photos, videos, UI screens, structural diagrams and creative materials are for reference only and may be cropped, compressed, composited or creatively rendered. Please refer to actual results and the final online page.",
    "zoom": "Zoom performance, equivalent focal lengths and telephoto results may vary depending on distance, lighting, mode, algorithm processing and device status. Optical-quality and digital zoom support is subject to the actual product.",
    "video": "Video, HDR, Log, LUT and motion photo specifications and effects depend on shooting mode, software version, storage, lighting and device status. Please refer to actual experience.",
    "software": "Software screens and features are shown for reference. AI and system capabilities may vary by software version, region, account login, network, app support and usage habits.",
    "battery": "Battery capacity, battery life and charging data are based on OPPO laboratory or supplier tests. Actual performance may vary depending on network, usage, charging conditions, device temperature, battery aging and other factors.",
    "durability": "Water, dust, drop and scratch resistance data is based on controlled laboratory conditions. Protection is not permanent and may decline with daily wear. Do not charge the device when wet; warranty coverage is subject to policy.",
    "performance": "Performance, benchmark, storage, thermal and gaming data are based on OPPO laboratory or official tests. Actual performance may vary by test environment, software version, memory configuration, network, app support and usage habits.",
    "connectivity": "Connectivity, audio, cross-device and third-party service performance may vary by region, carrier, network, device status, software version, app support and third-party service policies.",
}


RULES = [
    ("durability", r"IP\d|IPX|防水|防尘|抗摔|抗跌|抗划|金刚石|军用|SGS|water|dust|\bdrop\b|scratch|resistan"),
    ("battery", r"mAh|电池|续航|充电|闪充|无线充|反充|SUPERVOOC|AIRVOOC|battery|charging|charge"),
    ("video", r"8K|4K|\bLog\b|\bLUT\b|HDR|视频|电影|实况|Motion Photo|stabili[sz]ation|video"),
    ("zoom", r"变焦|长焦|远摄|焦段|mm|telephoto|zoom|focal|增距镜|optical-quality|digital zoom|10x|20x|120x|300mm"),
    ("camera", r"哈苏|镜头|影像|主摄|超广角|像素|大底|光圈|进光量|动态范围|丹霞|色彩|样张|拍摄|传感器|OIS|HNCS|camera|Hasselblad|lens|sensor|aperture|portrait|color|pixel|photo"),
    ("display", r"屏幕|亮度|刷新率|调光|护眼|低蓝光|边框|触控|nits|Hz|display|screen|brightness|refresh|eye"),
    ("software", r"AI|ColorOS|系统|界面|UI|软件|相册|助手|识别|抠图|流畅|OS|software|interface|album"),
    ("performance", r"芯片|平台|性能|跑分|内存|存储|散热|游戏|帧率|温度|CPU|GPU|RAM|storage|thermal|gaming|smooth"),
    ("connectivity", r"通信|网络|信号|NFC|互传|跨屏|音效|音量|第三方|运营商|NASA|connect|network|carrier|audio|third-party"),
    ("sample", r"样张|素材|CG|示意图|效果图|产品图|截图|动效|页面|render|sample|image|video material|diagram"),
    ("general", r"外观|设计|颜色|材质|工艺|尺寸|重量|厚度|宽度|皮革|中框|design|color|material|finish|leather"),
]


def norm(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"/", "／", "-", "—"}:
        return ""
    return re.sub(r"\s+", " ", text)


def row_text(ws, row: int, max_col: int) -> str:
    return " ".join(norm(ws.cell(row, col).value) for col in range(1, max_col + 1)).strip()


def is_english(text: str) -> bool:
    cjk = len(re.findall(r"[\u4e00-\u9fff]", text))
    letters = len(re.findall(r"[A-Za-z]", text))
    return letters > max(20, cjk * 2)


def find_header_row(ws) -> int | None:
    scan_rows = min(ws.max_row or 0, 12)
    for row in range(1, scan_rows + 1):
        values = {norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if values & HEADER_KEYS:
            return row
    return None


def existing_remark_cols(ws, header_row: int) -> list[int]:
    cols = []
    for col in range(1, ws.max_column + 1):
        header = norm(ws.cell(header_row, col).value).lower()
        if header in EXISTING_REMARK_HEADERS or header.replace(" ", "") in EXISTING_REMARK_HEADERS:
            cols.append(col)
    return cols


def valid_existing_note(text: str) -> str:
    text = norm(text)
    if not text or text in {"待补充", "TBD", "tbd"}:
        return ""
    if len(text) <= 2 and not re.search(r"\d|[A-Za-z\u4e00-\u9fff]", text):
        return ""
    return text


def categories_for(text: str) -> list[str]:
    cats = []
    for cat, pattern in RULES:
        if re.search(pattern, text, flags=re.IGNORECASE):
            cats.append(cat)
    if "camera" in cats and "sample" in cats:
        cats.remove("sample")
        cats.append("sample")
    if "zoom" in cats and "camera" not in cats:
        cats.append("camera")
    return cats[:3]


def merge_sentences(parts: list[str]) -> str:
    seen = set()
    output = []
    for part in parts:
        part = valid_existing_note(part)
        if not part:
            continue
        key = re.sub(r"\W+", "", part.lower())[:80]
        if key and key not in seen:
            seen.add(key)
            output.append(part)
    return "\n".join(output) if output else "/"


def generate_remark(text: str, existing: list[str]) -> str:
    if not text:
        return ""
    lang_templates = EN_TEMPLATES if is_english(text) else ZH_TEMPLATES
    cats = categories_for(text)
    generated = [lang_templates[cat] for cat in cats]
    return merge_sentences(existing + generated)


def copy_header_style(ws, source_col: int, target_col: int, header_row: int) -> None:
    for row in range(1, min(ws.max_row, header_row) + 1):
        src = ws.cell(row, source_col)
        dst = ws.cell(row, target_col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
    header = ws.cell(header_row, target_col)
    header.value = OUTPUT_HEADER
    header.font = copy(ws.cell(header_row, source_col).font) if ws.cell(header_row, source_col).has_style else Font(bold=True)
    header.fill = copy(ws.cell(header_row, source_col).fill) if ws.cell(header_row, source_col).has_style else PatternFill("solid", fgColor="D9EAD3")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def process_workbook(input_path: Path, output_path: Path) -> dict[str, int]:
    wb = load_workbook(input_path)
    stats: dict[str, int] = {}
    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if not header_row:
            continue
        existing_output = None
        for col in range(1, ws.max_column + 1):
            if norm(ws.cell(header_row, col).value) == OUTPUT_HEADER:
                existing_output = col
                break
        output_col = existing_output or ws.max_column + 1
        source_max_col = ws.max_column if existing_output is None else output_col - 1
        copy_header_style(ws, max(1, output_col - 1), output_col, header_row)
        remark_cols = [c for c in existing_remark_cols(ws, header_row) if c != output_col]
        written = 0
        for row in range(header_row + 1, ws.max_row + 1):
            text = row_text(ws, row, source_max_col)
            if not text:
                ws.cell(row, output_col).value = ""
                continue
            existing = [ws.cell(row, col).value for col in remark_cols]
            remark = generate_remark(text, existing)
            cell = ws.cell(row, output_col)
            cell.value = remark
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            written += 1
        ws.column_dimensions[get_column_letter(output_col)].width = 54
        stats[ws.title] = written
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    output = args.output or args.input.with_name(f"{args.input.stem}_备注生成.xlsx")
    stats = process_workbook(args.input, output)
    print(f"saved: {output}")
    for sheet, count in stats.items():
        print(f"{sheet}: {count} rows")


if __name__ == "__main__":
    main()
