#!/usr/bin/env python3
"""Generate sparse, risk-based official website remarks for workbook rows."""

from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_KEYS = {"章节", "卖点", "文案", "参数"}
EXISTING_REMARK_HEADERS = {"备注", "文案备注", "备注说明", "remark", "remarks", "note", "notes"}
OUTPUT_HEADER = "备注生成"
SKIP_HEADERS = {"素材", "素材确认", "素材 / 样张需求", "建议", "友商", "相机厂商", "文案逻辑"}
SKIP_HEADER_KEYWORDS = {"素材", "样张变化", "样张需求", "参考图", "建议", "友商", "相机厂商"}

GLOBAL_ZH = (
    "* 如无特殊备注，以上页面中所示数据均为 OPPO 实验室数据，仅供参考。实际数据受使用环境、设备状况、"
    "软件版本等因素影响将略有差异，请以实际数据为准。\n"
    "* 如无特殊备注，以上页面中所示产品界面内容（包括但不限于 UI、背景）仅作功能展示，各个功能分属不同"
    "软件版本，不同软件版本的上线时间与区域有所差异，请以实际版本功能及界面为准。AI 相关能力因所处环境、"
    "网络环境、个人使用习惯等因素不同，实际体验效果可能不同，请以实际体验为准。\n"
    "* 产品图仅供参考，请以实物为准。"
)

GLOBAL_EN = (
    "* Unless otherwise stated, data shown on this page is based on OPPO laboratory tests and is for reference only. "
    "Actual results may vary depending on environment, device status, software version and other factors.\n"
    "* Product interfaces and software features shown on this page are for reference only. Availability may vary by "
    "software version, region, network, account status and app support.\n"
    "* Product images are for reference only. Please refer to the actual product."
)

ZH_TEMPLATES = {
    "availability": "相关功能的支持范围、上线时间、适配应用、账号或网络要求可能因地区、软件版本及服务策略不同而变化，请以实际支持情况为准。",
    "ai": "AI 相关功能的生成内容仅供参考，实际效果可能因所处环境、网络环境、识别对象、语种及个人使用习惯不同而存在差异。",
    "accessory": "相关配件或套装需单独购买，具体配置、兼容范围和销售信息请以官方网站介绍为准。",
    "waterproof": "防水、防尘、抗摔、抗划等能力基于受控实验室条件，防护能力并非永久有效，可能因日常磨损而下降；请勿在潮湿状态下充电。",
    "battery": "电池容量、续航、充电功率及充电时间为特定测试条件下数据，实际表现可能因使用方式、充电条件、设备温度、电池老化等因素不同而存在差异。",
    "display_eye": "护眼相关功能非医疗器械功能，不具有治疗作用；认证、亮度、刷新率等信息请以实际证书、产品规格和使用体验为准。",
    "camera_mode": "相关拍摄规格或影像效果需在指定模式、焦段或功能开关下生效，实际效果可能因拍摄环境、方法、设备状态和软件版本不同而存在差异。",
    "video_sample": "页面样张、样片及视频效果仅作功能演示，可能经过裁切、压缩或创意化呈现，请以实际拍摄效果为准。",
    "satellite": "卫星通信、网络通信及运营商相关能力受地区、运营商开通规则、网络覆盖、设备版本及软件升级影响，请以实际支持情况为准。",
    "certification": "认证信息以实际证书和官方发布信息为准，认证效果不代表医疗、绝对防护或永久有效承诺。",
}

EN_TEMPLATES = {
    "availability": "Feature availability, launch timing, app support, account or network requirements may vary by region, software version and service policy. Please refer to actual support.",
    "ai": "AI-generated content is for reference only. Actual results may vary depending on environment, network, subject, language and usage habits.",
    "accessory": "Related accessories or kits are sold separately. Configuration, compatibility and sales information are subject to the official website.",
    "waterproof": "Water, dust, drop and scratch resistance are based on controlled laboratory conditions and are not permanent. Protection may decrease with daily wear. Do not charge the device when wet.",
    "battery": "Battery capacity, battery life, charging power and charging time are based on specific test conditions. Actual performance may vary by usage, charging conditions, device temperature and battery aging.",
    "display_eye": "Eye-care features are not medical functions. Certification, brightness and refresh rate information are subject to actual certificates, product specifications and experience.",
    "camera_mode": "Camera specifications or imaging effects may require specific modes, focal lengths or feature switches. Actual results may vary by shooting environment, method, device status and software version.",
    "video_sample": "Sample photos and videos are for feature demonstration only and may be cropped, compressed or creatively rendered. Please refer to actual shooting results.",
    "satellite": "Satellite, network and carrier-related capabilities depend on region, carrier rules, coverage, device version and software updates. Please refer to actual support.",
    "certification": "Certification information is subject to actual certificates and official information. Certification does not imply medical effect, absolute protection or permanent performance.",
}

RISK_RULES = [
    ("waterproof", r"IP\d|IPX|防水|防尘|抗摔|抗跌|抗划|耐摔|金刚石|军用|五星整机|water|dust|\bdrop\b|scratch|resistan"),
    ("satellite", r"卫星|天通卫星|北斗|运营商|中国移动|中国联通|中国电信|n79|carrier|satellite"),
    ("accessory", r"单独购买|另售|套装|配件|增距镜|手柄|滤镜|NAS|accessor|kit|sold separately|teleconverter"),
    ("ai", r"AI\s*(一键|生成|消除|翻译|菜单|旅行|灵感|闪记|问屏|补光|能力)|小布|一键闪记|智能生成|生成内容|AI\s*generated"),
    ("availability", r"后续|OTA|陆续适配|具体适配|适配应用|仅支持|仅.*支持|部分应用|具体支持|上线时间|需在|需升级|需安装|需登录|需要.*(?:升级|安装|登录|开启|购买)|账号|登录|网络环境|地区|应用版本|软件版本|尝鲜版|support.*app|software version|region|account|network"),
    ("certification", r"获得[^。；\n]*认证|证书编号|TÜV|莱茵|SGS|certificat"),
    ("battery", r"\d+\s*mAh|Wh|电池容量|续航|充电|闪充|无线充|循环寿命|典型值|额定值|SUPERVOOC|AIRVOOC|battery|charging|charge"),
    ("display_eye", r"护眼|低蓝光|无频闪|1nit|亮度|刷新率|PWM|eye-care|low blue|brightness|refresh rate"),
    ("camera_mode", r"哈苏超清模式|大师模式|打开.*开关|变焦|长焦|超清照片|实况|\bLog\b|\bLUT\b|HDR|OIS|防抖|zoom|focal|telephoto|motion photo"),
    ("video_sample", r"样片|视频样片|仅作功能演示|拍摄效果|sample photo|sample video|for demonstration"),
    ("comparison", r"对比|相比|提升|降低|减少|高于|优于|行业|首发|首个|首款|首次|唯一|最大|最高|最强|最薄|更高|更低|更长|更短|更少|更快|领先|首创|独家|峰值|至高|高达|\d+\s*倍|up to|first|only|leading|increase|decrease|compared"),
]


def norm(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"/", "／", "-", "—"}:
        return ""
    return re.sub(r"\s+", " ", text)


def is_english(text: str) -> bool:
    cjk = len(re.findall(r"[\u4e00-\u9fff]", text))
    letters = len(re.findall(r"[A-Za-z]", text))
    return letters > max(20, cjk * 2)


def find_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row or 0, 12) + 1):
        values = {norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if values & HEADER_KEYS:
            return row
    return None


def header_value(ws, header_row: int, col: int) -> str:
    return norm(ws.cell(header_row, col).value)


def is_skipped_header(header: str) -> bool:
    return header in SKIP_HEADERS or any(keyword in header for keyword in SKIP_HEADER_KEYWORDS)


def existing_remark_cols(ws, header_row: int) -> list[int]:
    cols = []
    for col in range(1, ws.max_column + 1):
        header = header_value(ws, header_row, col).lower()
        if header in EXISTING_REMARK_HEADERS or header.replace(" ", "") in EXISTING_REMARK_HEADERS:
            cols.append(col)
    return cols


def screen_col(ws, header_row: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        if header_value(ws, header_row, col) == "屏数":
            return col
    return None


def public_row_text(ws, row: int, header_row: int, max_col: int) -> str:
    parts = []
    for col in range(1, max_col + 1):
        header = header_value(ws, header_row, col)
        if not header or is_skipped_header(header):
            continue
        parts.append(norm(ws.cell(row, col).value))
    return " ".join(part for part in parts if part).strip()


def valid_existing_note(text) -> str:
    text = norm(text)
    if not text or text in {"待补充", "TBD", "tbd"}:
        return ""
    if len(text) <= 2 and not re.search(r"\d|[A-Za-z\u4e00-\u9fff]", text):
        return ""
    return text


def curated_remark_count(ws, header_row: int, remark_cols: list[int]) -> int:
    count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        for col in remark_cols:
            note = valid_existing_note(ws.cell(row, col).value)
            if note:
                count += 1
                break
    return count


def looks_like_global_note(text: str) -> bool:
    return bool(re.search(r"如无特殊备注|未经特殊说明|unless otherwise stated", text, re.I))


def looks_like_overview_list(text: str) -> bool:
    head = text[:160]
    return bool(re.search(r"Highlights|KSP|概览", head, re.I))


def has_footnote_marker(text: str) -> bool:
    return bool(re.search(r"[（(]\s*\d+\s*[）)]|\*\s*$", text))


def has_camera_condition(text: str) -> bool:
    return bool(
        re.search(
            r"哈苏超清模式|哈苏大师模式|大师模式|打开.*开关|须关闭|须在|需在.*模式|仅支持.*焦段|\bLog\b|\bLUT\b|"
            r"瞬时三曝光|三次\s*AA|所有镜头均全新设计|微距|4K.*实况|8K|全焦段",
            text,
            flags=re.IGNORECASE,
        )
    )


def has_battery_condition(text: str) -> bool:
    return bool(
        re.search(r"\d+\s*mAh|Wh|电池容量|续航|充电|闪充|无线充|循环寿命|单电芯|典型值|额定值|低温.*续航", text, re.I)
        and re.search(r"提升|对比|相比|循环|单电芯|典型值|额定值|充电功率|无线充功率|充电时间|低温|收益|mAh|Wh", text, re.I)
    )


def risk_categories(text: str) -> list[str]:
    if not text:
        return []
    cats = []
    for cat, pattern in RISK_RULES:
        if re.search(pattern, text, flags=re.IGNORECASE):
            cats.append(cat)

    comparison_allowed = "comparison" in cats and measurement_note_zh(text) is not None
    camera_allowed = "camera_mode" in cats and has_camera_condition(text)
    # Pure parameters and broad "highest/first/leading" claims are covered by global notes
    # or require copy/legal revision, not a vague per-row remark.
    high_risk = {"availability", "ai", "accessory", "waterproof", "satellite", "certification", "video_sample"}
    if not has_footnote_marker(text) and not (set(cats) & high_risk) and not comparison_allowed and not camera_allowed:
        return []

    priority = [
        "comparison",
        "availability",
        "ai",
        "accessory",
        "waterproof",
        "satellite",
        "certification",
        "battery",
        "display_eye",
        "camera_mode",
        "video_sample",
    ]
    ordered = []
    for cat in priority:
        if cat not in cats:
            continue
        if cat == "comparison" and not comparison_allowed:
            continue
        if cat == "camera_mode" and not (camera_allowed or has_footnote_marker(text)):
            continue
        if cat == "battery" and ("accessory" in cats or not has_battery_condition(text)):
            continue
        if cat == "display_eye" and "certification" in cats:
            continue
        if cat == "certification" and ("waterproof" in cats or len(text.strip()) < 20):
            continue
        if cat == "video_sample" and "接样片" not in text and "样片" not in text and "仅作功能演示" not in text and "拍摄效果" not in text:
            continue
        ordered.append(cat)
    return ordered[:2]


def merge_notes(parts: list[str]) -> str:
    seen = set()
    output = []
    for part in parts:
        part = valid_existing_note(part)
        if not part:
            continue
        key = re.sub(r"\W+", "", part.lower())[:80]
        if key and key not in seen:
            seen.add(key)
            output.append(part)
    return "\n".join(output) if output else "/"


def split_note_lines(remark: str) -> list[str]:
    if not remark or remark == "/":
        return []
    return [line.strip() for line in str(remark).splitlines() if line.strip() and line.strip() != "/"]


def dedupe_within_screen(remark: str, screen_key: str, seen_by_screen: dict[str, set[str]]) -> str:
    if not screen_key or looks_like_global_note(remark):
        return remark
    seen = seen_by_screen.setdefault(screen_key, set())
    kept = []
    for line in split_note_lines(remark):
        key = re.sub(r"\W+", "", line.lower())[:120]
        if key in seen:
            continue
        seen.add(key)
        kept.append(line)
    return "\n".join(kept) if kept else "/"


def compact_claim(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", text).strip()
    clauses = re.split(r"[。；;\n]", cleaned)
    claim_words = r"对比|相比|提升|降低|高于|优于|行业|首发|首个|首款|首次|唯一|最大|最高|最强|最薄|更高|更低|更长|更短|更少|更快|领先|首创|独家|峰值|至高|高达|up to|first|only|leading|increase|decrease|compared"
    for clause in clauses:
        clause = clause.strip(" ，,：:")
        match = re.search(claim_words, clause, re.I)
        if 6 <= len(clause) <= 58 and match:
            return clause
        if 6 <= len(clause) <= 80 and match:
            start = match.start()
            if not re.match(r"行业|首发|首个|首款|首次|唯一|最大|最高|最强|最薄|领先|首创|独家|至高|高达", match.group(0), re.I) and start > 16:
                start = max(0, start - 8)
            return clause[start : start + 58].strip(" ，,：:")
    match = re.search(r"([\u4e00-\u9fffA-Za-z0-9 /+×&.\"'%-]{6,46}(?:提升|降低|高于|优于|行业|首发|首个|首款|首次|唯一|最大|最高|领先|首创|独家|峰值|至高|高达)[\u4e00-\u9fffA-Za-z0-9 /+×&.\"'%-]{0,28})", cleaned, re.I)
    if match:
        return match.group(1).strip(" ，,：:")
    return cleaned[:36].strip(" ，,：:")


def measurement_note_zh(text: str) -> str | None:
    variation = "因测试环境、方法、软件版本、设备状态等不同，数据可能存在差异，请以产品实际为准。"
    if "进光量" in text:
        if not re.search(r"对比|相比|上一代|一英寸|S23U|提升(?:至|到)?\s*\d", text, re.I):
            return None
        if "主摄" in text:
            subject = "主摄镜头进光量"
        elif "长焦" in text:
            subject = "长焦镜头进光量"
        elif "超广角" in text:
            subject = "超广角镜头进光量"
        else:
            subject = "镜头进光量"
        return f"{subject}数据来自 OPPO 实验室，相关提升或对比结果基于页面所述测试对象及测试条件，{variation}"
    if "感光效率" in text and re.search(r"提升|高于|对比|相比", text):
        return f"感光效率数据来自 OPPO 实验室，相关提升结果基于特定测试条件，{variation}"
    if "像素" in text and re.search(r"像素.*提升|提升.*像素|高于|对比|相比", text):
        return f"像素或画质提升数据来自 OPPO 实验室，相关结果基于页面所述测试对象及测试条件，{variation}"
    if re.search(r"动态范围|宽容度|动态提升", text):
        return f"动态范围相关数据来自 OPPO 实验室，相关提升结果基于特定测试条件，{variation}"
    if re.search(r"GPU|CPU|功耗|内存占用|骁龙|跑分", text, re.I):
        return "性能、功耗、内存占用或芯片相关数据来自 OPPO 实验室或芯片平台官方信息，实际表现可能因测试环境、软件版本、设备状态、应用适配等不同而存在差异。"
    if re.search(r"续航|充电|mAh|Wh|充电功率|无线充功率|闪充", text, re.I) and re.search(r"提升|对比|相比|循环|单电芯|典型值|额定值|充电时间|低温|收益", text):
        return "电池容量、续航和充电相关数据来自 OPPO 实验室，实际表现可能因网络环境、使用方式、充电条件、设备温度、电池老化等因素不同而存在差异，请以实际体验为准。"
    if re.search(r"防抖|OIS|抖动|校准|AA", text, re.I) and re.search(r"最大|提升|三次|对比|相比", text):
        return "防抖、光学校准或稳定性相关数据来自 OPPO 实验室，实际效果可能因拍摄环境、设备状态、软件版本及测试方法不同而存在差异，请以实际体验为准。"
    return None


def comparison_note_en(text: str) -> str:
    measured = measurement_note_zh(text)
    if measured:
        return measured
    return ""


def comparison_note(text: str) -> str:
    if is_english(text):
        return comparison_note_en(text)
    measured = measurement_note_zh(text)
    if measured:
        return measured
    return ""


def category_note_zh(cat: str, text: str) -> str:
    if cat == "availability":
        if "一键闪记" in text:
            return "AI 一键闪记相关参数或部分功能体验仅部分应用支持，具体支持的应用、功能和界面，请以实际为准。AI 生成内容仅供参考。"
        if "小布记忆" in text or "旅行合集" in text:
            return "小布记忆相关 AI 体验的部分能力需后续软件 OTA 升级支持。AI 生成内容仅供参考，具体功能与界面请以实际版本为准。"
        if "菜单翻译" in text:
            return "AI 菜单翻译支持语种及识别效果可能因语种、所处环境、网络环境、个人使用习惯等因素不同而变化。AI 生成内容仅供参考，请以实际体验为准。"
        if "到站提醒" in text:
            return "小布建议智能到站提醒的支持城市和路线将随版本陆续适配，请以官方信息和实际支持情况为准。"
        if "NAS" in text:
            return "该功能需后续软件 OTA 升级支持，所需 NAS 个人存储需单独购买。"
        if re.search(r"iOS|iPhone|跨生态|互传|OPPO 互联|O\+ 互联", text, re.I):
            return "与 iOS 设备跨生态互传时，需在 iPhone 端安装并升级相关 OPPO 互联应用，具体支持情况请以实际为准。"
        if "音频共享" in text or "A2DP" in text:
            return "非 A2DP 类型耳机不支持音频共享，具体支持情况请以实际为准。"
        if "锁屏沉浸" in text or "导航锁屏" in text:
            return "锁屏沉浸态及导航、音乐等锁屏显示能力受适配应用和应用版本影响，请将相关应用升级至最新版本，具体适配以实际为准。"
    if cat == "accessory":
        if "增距镜" in text and ("套装" in text or "摄影" in text):
            return "相关摄影配件或套装需要单独购买，套装内容、兼容范围和销售信息请以官方网站介绍为准；滤镜、转接环等配件请注意实际适配尺寸。"
        return "需要单独购买，详细信息以官方网站介绍为准。"
    if cat == "ai":
        if "菜单翻译" in text:
            return "AI 菜单翻译结果仅供参考，实际体验会因语种、所处环境、网络环境及个人使用习惯等因素不同而变化。"
        if "一键闪记" in text:
            return "AI 一键闪记部分功能体验仅部分应用支持，具体支持的应用、功能和界面，请以实际为准。AI 生成内容仅供参考。"
        return ZH_TEMPLATES[cat]
    if cat == "camera_mode":
        if re.search(r"哈苏大师视频|大师视频|\bLog\b", text, re.I):
            return "使用哈苏大师视频影调进行专业视频录制时，须关闭 Log。视频规格和效果可能因拍摄模式、软件版本、存储空间及设备状态不同而存在差异。"
        if "所有镜头均全新设计" in text:
            return "“所有镜头均全新设计”是指本机型包含的所有光学镜头相对于既往发布的 OPPO 系列手机镜头均为全新设计，具体镜头范围以页面列示为准。"
        if "三次 AA" in text:
            return "“三次 AA”是指经过三次超精密主动光学校准，以实现更高的光学精度。"
        if "瞬时三曝光" in text:
            return "“瞬时三曝光”为算法名，指传感器在拍摄时读出三次不同曝光信息，以获得更好的画面动态和抓拍效果；实际效果可能因拍摄环境、方法、设备状态等不同而存在差异。"
        if "微距" in text:
            return "长焦微距对焦距离数据来自 OPPO 实验室，因测试环境、方法、软件版本等不同，数据可能存在差异。"
        if "2 亿像素" in text and "照片" in text:
            return "2 亿像素照片需在指定拍摄模式下开启对应像素开关，支持焦段和实际效果请以产品实际为准。"
        if "4K" in text and "实况" in text:
            return "4K 超清实况功能需在指定拍摄模式下开启对应功能开关，实际效果可能因拍摄环境、方法、设备状态等不同而存在差异，请以产品实际为准。"
        if "全焦段" in text or "8K" in text:
            return "全焦段、8K 或高像素画质相关能力需在产品支持的指定焦段、模式或格式下生效，实际效果可能因拍摄环境、方法、设备状态等不同而存在差异，请以产品实际为准。"
        if "丹霞" in text:
            return "丹霞色彩还原因拍摄环境、方法、设备状态等不同，可能存在差异，请以产品实际为准。"
    if cat == "video_sample":
        if "视频" in text:
            return "视频样片仅作功能演示，实际视频拍摄效果可能因拍摄环境、方法、设备状态和软件版本不同而存在差异，请以实际拍摄效果为准。"
        return ZH_TEMPLATES[cat]
    if cat == "satellite":
        return "卫星通信功能仅对应版本支持，天通、北斗及运营商服务范围、开通规则和上线时间可能因运营商、地区及软件升级不同而变化，请以实际支持情况和运营商信息为准。"
    if cat == "waterproof":
        return "产品在正常使用状态下可防水，相关 IP 防护等级基于受控实验室条件；生活场景与实验环境存在差异，请勿在潮湿状态下充电，防水、防尘能力可能因日常使用损耗而下降。"
    if cat == "certification":
        if re.search(r"护眼|低蓝光|无频闪|TÜV|莱茵", text, re.I):
            return "护眼、低蓝光、无频闪等认证信息以实际证书为准；本产品非医疗器械，不具备治疗能力。"
        return ZH_TEMPLATES[cat]
    return ZH_TEMPLATES.get(cat, "")


def generate_remark(text: str, existing: list[str], *, is_last_content_row: bool) -> str:
    existing_notes = [valid_existing_note(note) for note in existing]
    existing_notes = [note for note in existing_notes if note]
    if existing_notes:
        return merge_notes(existing_notes)

    if looks_like_global_note(text):
        return text

    cats = risk_categories(text)
    if not cats:
        return GLOBAL_EN if is_last_content_row and is_english(text) else (GLOBAL_ZH if is_last_content_row else "/")

    parts = []
    english = is_english(text)
    for cat in cats:
        if cat == "comparison":
            parts.append(comparison_note(text))
        elif english:
            parts.append(EN_TEMPLATES[cat])
        else:
            parts.append(category_note_zh(cat, text))
    return merge_notes(parts)


def copy_header_style(ws, source_col: int, target_col: int, header_row: int) -> None:
    for row in range(1, min(ws.max_row, header_row) + 1):
        src = ws.cell(row, source_col)
        dst = ws.cell(row, target_col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
    header = ws.cell(header_row, target_col)
    header.value = OUTPUT_HEADER
    header.font = copy(ws.cell(header_row, source_col).font) if ws.cell(header_row, source_col).has_style else Font(bold=True)
    header.fill = copy(ws.cell(header_row, source_col).fill) if ws.cell(header_row, source_col).has_style else PatternFill("solid", fgColor="D9EAD3")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def last_content_row(ws, header_row: int, source_max_col: int) -> int:
    for row in range(ws.max_row, header_row, -1):
        if public_row_text(ws, row, header_row, source_max_col):
            return row
    return ws.max_row


def process_workbook(input_path: Path, output_path: Path) -> dict[str, dict[str, int]]:
    wb = load_workbook(input_path)
    stats: dict[str, dict[str, int]] = {}
    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if not header_row:
            continue
        existing_output = None
        for col in range(1, ws.max_column + 1):
            if norm(ws.cell(header_row, col).value) == OUTPUT_HEADER:
                existing_output = col
                break
        output_col = existing_output or ws.max_column + 1
        source_max_col = ws.max_column if existing_output is None else output_col - 1
        copy_header_style(ws, max(1, output_col - 1), output_col, header_row)
        remark_cols = [c for c in existing_remark_cols(ws, header_row) if c != output_col]
        curated_only = curated_remark_count(ws, header_row, remark_cols) >= 8
        final_row = last_content_row(ws, header_row, source_max_col)
        page_col = screen_col(ws, header_row)
        last_screen = ""
        seen_by_screen: dict[str, set[str]] = {}
        rows_seen = 0
        remarks_written = 0
        for row in range(header_row + 1, ws.max_row + 1):
            text = public_row_text(ws, row, header_row, source_max_col)
            if not text:
                ws.cell(row, output_col).value = ""
                continue
            if page_col:
                current_screen = norm(ws.cell(row, page_col).value)
                if current_screen:
                    last_screen = current_screen
                screen_key = last_screen
            else:
                screen_key = ""
            existing = [ws.cell(row, col).value for col in remark_cols]
            has_existing_note = any(valid_existing_note(note) for note in existing)
            if looks_like_overview_list(text) and not has_existing_note and row != final_row:
                remark = "/"
            elif curated_only and not has_existing_note:
                remark = GLOBAL_EN if row == final_row and is_english(text) else (GLOBAL_ZH if row == final_row else "/")
            else:
                remark = generate_remark(text, existing, is_last_content_row=(row == final_row))
            if not has_existing_note:
                remark = dedupe_within_screen(remark, screen_key, seen_by_screen)
            cell = ws.cell(row, output_col)
            cell.value = remark
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            rows_seen += 1
            if remark and remark != "/":
                remarks_written += 1
        ws.column_dimensions[get_column_letter(output_col)].width = 54
        stats[ws.title] = {"rows": rows_seen, "remarks": remarks_written}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    output = args.output or args.input.with_name(f"{args.input.stem}_备注生成.xlsx")
    stats = process_workbook(args.input, output)
    print(f"saved: {output}")
    for sheet, sheet_stats in stats.items():
        print(f"{sheet}: {sheet_stats['remarks']} remarks / {sheet_stats['rows']} content rows")


if __name__ == "__main__":
    main()
